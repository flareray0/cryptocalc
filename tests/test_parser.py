from __future__ import annotations

from decimal import Decimal
import re
from zipfile import ZipFile
from pathlib import Path

from openpyxl import Workbook

from app.parsers.binance_japan_parser import BinanceJapanParser


def _force_sheet_dimension_to_a1(path: Path) -> None:
    temp_path = path.with_suffix(".tmp.xlsx")
    with ZipFile(path, "r") as src, ZipFile(temp_path, "w") as dst:
        for entry in src.infolist():
            payload = src.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                text = payload.decode("utf-8")
                text = re.sub(r'<dimension ref="[^"]+"', '<dimension ref="A1"', text, count=1)
                payload = text.encode("utf-8")
            dst.writestr(entry, payload)
    temp_path.replace(path)


def test_binance_japan_parser_reads_sample():
    parser = BinanceJapanParser()
    sample = Path(r"H:\cryptocalc\samples\binance_japan_sample.csv")
    batch = parser.parse(sample)

    assert batch.transaction_count == 4
    assert batch.transactions[0].base_asset == "BTC"
    assert batch.transactions[0].gross_amount_jpy is not None
    assert batch.transactions[2].tx_type.value == "crypto_swap"
    assert batch.transactions[3].review_flag is True


def test_binance_japan_parser_reads_japanese_balance_history(tmp_path):
    sample = tmp_path / "binance_jp_balance_history.csv"
    sample.write_text(
        "\n".join(
            [
                "ユーザーID,時間,アカウント,操作,コイン,変更,備考",
                "1,25-07-17 09:18:07,Spot,Deposit,JPY,15000,",
                "1,25-07-17 09:55:49,Spot,Binance Convert,JPY,-15000,",
                "1,25-07-17 09:55:49,Spot,Binance Convert,ETH,0.02929314,",
                "1,25-07-17 10:37:15,Spot,Transfer Between Main and Funding Wallet,BTC,-0.0002,",
                "1,25-07-17 10:37:15,Funding,Transfer Between Main and Funding Wallet,BTC,0.0002,",
                "1,25-07-17 11:06:11,Spot,Transaction Fee,ETH,-0.0000035,",
                "1,25-07-17 11:06:11,Spot,Transaction Buy,ETH,0.0035,",
                "1,25-07-17 11:06:11,Spot,Transaction Spend,BTC,-0.00009908,",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    parser = BinanceJapanParser()
    batch = parser.parse(sample)

    assert batch.transaction_count == 3
    assert batch.unknown_column_names == []
    assert batch.unknown_tx_types == []
    assert batch.transactions[0].tx_type.value == "transfer_in"
    assert batch.transactions[0].base_asset == "JPY"
    assert batch.transactions[1].tx_type.value == "buy"
    assert batch.transactions[1].base_asset == "ETH"
    assert batch.transactions[1].gross_amount_jpy == 15000
    assert batch.transactions[2].tx_type.value == "crypto_swap"
    assert batch.transactions[2].base_asset == "ETH"
    assert batch.transactions[2].quote_asset == "BTC"


def test_binance_japan_parser_reads_report_style_xlsx_with_offset_header(tmp_path):
    sample = tmp_path / "binance_jp_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"

    ws["C3"] = "取引履歴"
    ws["C5"] = "名前"
    ws["D5"] = "匿名ユーザー"
    ws["F5"] = "メール"
    ws["G5"] = "masked@example.com"
    ws["C6"] = "ユーザーID"
    ws["D6"] = "123456"
    ws["F6"] = "期間(UTC+9)"
    ws["G6"] = "2026-03-14 to 2026-03-21"

    headers = {
        "C10": "ユーザーID",
        "D10": "時間",
        "F10": "アカウント",
        "G10": "操作",
        "I10": "コイン",
        "J10": "変更",
        "L10": "備考",
    }
    for cell, value in headers.items():
        ws[cell] = value

    rows = [
        ("C11", "123456"), ("D11", "26-03-15 18:27:19"), ("F11", "Spot"), ("G11", "Transaction Sold"), ("I11", "BTC"), ("J11", "-0.002285"), ("L11", ""),
        ("C12", "123456"), ("D12", "26-03-15 18:27:19"), ("F12", "Spot"), ("G12", "Transaction Revenue"), ("I12", "JPY"), ("J12", "26256.950995"), ("L12", ""),
        ("C13", "123456"), ("D13", "26-03-15 18:27:19"), ("F13", "Spot"), ("G13", "Transaction Fee"), ("I13", "JPY"), ("J13", "-11.5346793"), ("L13", ""),
        ("C14", "123456"), ("D14", "26-03-16 10:00:00"), ("F14", "Spot"), ("G14", "Deposit"), ("I14", "JPY"), ("J14", "15000"), ("L14", ""),
    ]
    for cell, value in rows:
        ws[cell] = value

    wb.save(sample)
    _force_sheet_dimension_to_a1(sample)

    parser = BinanceJapanParser()
    batch = parser.parse(sample)

    assert batch.detected_layout == "xlsx_japanese_balance_history_report"
    assert batch.header_row_number == 10
    assert batch.transaction_count == 2
    assert batch.unknown_column_names == []
    assert batch.unknown_tx_types == []
    assert batch.transactions[0].tx_type.value == "sell"
    assert batch.transactions[0].gross_amount_jpy == Decimal("26256.950995")
    assert batch.transactions[1].tx_type.value == "transfer_in"
    assert batch.transactions[1].base_asset == "JPY"


def test_binance_japan_parser_reads_trade_export_xlsx(tmp_path):
    sample = tmp_path / "binance_trade_export.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(
        [
            "Date(UTC)",
            "Pair",
            "Base Asset",
            "Quote Asset",
            "Type",
            "Price",
            "Amount",
            "Total",
            "Fee",
            "Fee Coin",
        ]
    )
    ws.append(
        [
            "2025-01-03 00:00:00",
            "BTCJPY",
            "BTC",
            "JPY",
            "BUY",
            "10000000",
            "0.001",
            "10000",
            "10",
            "JPY",
        ]
    )
    wb.save(sample)

    parser = BinanceJapanParser()
    batch = parser.parse(sample)

    assert batch.detected_layout == "xlsx_trade_export"
    assert batch.header_row_number == 1
    assert batch.transaction_count == 1
    assert batch.transactions[0].tx_type.value == "buy"
    assert batch.transactions[0].gross_amount_jpy == Decimal("10000")


def test_binance_japan_parser_reads_spot_trade_history_csv(tmp_path):
    sample = tmp_path / "binance_spot_trade_history.csv"
    sample.write_text(
        "\n".join(
            [
                "時間,ペア,サイド,価格,実行済み,金額,手数料",
                "26-03-20 23:39:35,BTCJPY,BUY,10000000,0.001BTC,10000JPY,10JPY",
                "26-03-20 23:31:58,SOLETH,SELL,0.04181,0.169SOL,0.00706589ETH,0.00000565ETH",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    batch = BinanceJapanParser().parse(sample)

    assert batch.detected_layout == "csv_spot_trade_history"
    assert batch.transaction_count == 2
    assert batch.unknown_column_names == []
    assert batch.transactions[0].tx_type.value == "buy"
    assert batch.transactions[0].gross_amount_jpy == Decimal("10000")
    assert batch.transactions[0].fee_jpy == Decimal("10")
    assert batch.transactions[1].tx_type.value == "crypto_swap"
    assert batch.transactions[1].base_asset == "SOL"
    assert batch.transactions[1].quote_asset == "ETH"
    assert batch.transactions[1].review_flag is True


def test_binance_japan_parser_reads_empty_deposit_and_withdraw_history_csv(tmp_path):
    deposit = tmp_path / "binance_deposit_history.csv"
    deposit.write_text(
        "\n".join(
            [
                "時間,コイン,ネットワーク,金額,住所,トランザクションID,ステータス",
                "条件に一致するデータがありません。",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    withdraw = tmp_path / "binance_withdraw_history.csv"
    withdraw.write_text(
        "\n".join(
            [
                "時間,コイン,ネットワーク,金額,手数料,住所,トランザクションID,ステータス",
                "条件に一致するデータがありません。",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    deposit_batch = BinanceJapanParser().parse(deposit)
    withdraw_batch = BinanceJapanParser().parse(withdraw)

    assert deposit_batch.detected_layout == "csv_deposit_history"
    assert deposit_batch.transaction_count == 0
    assert deposit_batch.review_required_count == 0
    assert withdraw_batch.detected_layout == "csv_withdraw_history"
    assert withdraw_batch.transaction_count == 0
    assert withdraw_batch.review_required_count == 0


def test_binance_japan_parser_reads_fiat_deposit_and_withdraw_history_csv(tmp_path):
    deposit = tmp_path / "binance_fiat_deposit_history.csv"
    deposit.write_text(
        "\n".join(
            [
                "時間,方法,入金額,受取金額,手数料,ステータス,取引ID",
                "26-03-11 18:22:34,Bank Transfer (GMO),17198.00 JPY,17198 JPY,0.00 JPY,Successful,OF1",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    withdraw = tmp_path / "binance_fiat_withdraw_history.csv"
    withdraw.write_text(
        "\n".join(
            [
                "時間,方法,出金額,受取金額,手数料,ステータス,取引ID",
                "26-03-11 18:17:16,Bank Transfer (GMO),102000.00 JPY,101850 JPY,150.00 JPY,Successful,WF1",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    deposit_batch = BinanceJapanParser().parse(deposit)
    withdraw_batch = BinanceJapanParser().parse(withdraw)

    assert deposit_batch.detected_layout == "csv_fiat_deposit_history"
    assert deposit_batch.transaction_count == 1
    assert deposit_batch.transactions[0].tx_type.value == "transfer_in"
    assert deposit_batch.transactions[0].gross_amount_jpy == Decimal("17198")
    assert withdraw_batch.detected_layout == "csv_fiat_withdraw_history"
    assert withdraw_batch.transaction_count == 1
    assert withdraw_batch.transactions[0].tx_type.value == "transfer_out"
    assert withdraw_batch.transactions[0].gross_amount_jpy == Decimal("102000.00")
    assert withdraw_batch.transactions[0].fee_jpy == Decimal("150.00")


def test_binance_japan_parser_reads_empty_fiat_conversion_history_csv(tmp_path):
    conversion = tmp_path / "binance_fiat_conversion_history.csv"
    conversion.write_text(
        "\n".join(
            [
                "方法,金額,価格,最終金額,作成日時,ステータス,取引ID",
                "条件に一致するデータがありません。",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    batch = BinanceJapanParser().parse(conversion)

    assert batch.detected_layout == "csv_fiat_conversion_history"
    assert batch.transaction_count == 0
    assert batch.review_required_count == 0

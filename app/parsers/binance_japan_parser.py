from __future__ import annotations

import csv
import hashlib
import re
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

from app.domain.enums import ClassificationStatus, ImportSourceKind, Side, TransactionType
from app.domain.models import ImportBatchResult, NormalizedTransaction
from app.domain.validators import JST, parse_utc_timestamp, safe_slug, to_decimal, utc_to_jst
from app.parsers.base_parser import BaseParser

EXPECTED_COLUMNS = [
    "Date(UTC)",
    "Pair",
    "Base Asset",
    "Quote Asset",
    "Type",
    "Price",
    "Amount",
    "Total",
    "Fee",
    "Fee Coin",
]

EXPECTED_COLUMNS_JP = [
    "ユーザーID",
    "時間",
    "アカウント",
    "操作",
    "コイン",
    "変更",
    "備考",
]

EXPECTED_COLUMNS_SPOT_TRADE_JP = [
    "時間",
    "ペア",
    "サイド",
    "価格",
    "実行済み",
    "金額",
    "手数料",
]

EXPECTED_COLUMNS_DEPOSIT_JP = [
    "時間",
    "コイン",
    "ネットワーク",
    "金額",
    "住所",
    "トランザクションID",
    "ステータス",
]

EXPECTED_COLUMNS_WITHDRAW_JP = [
    "時間",
    "コイン",
    "ネットワーク",
    "金額",
    "手数料",
    "住所",
    "トランザクションID",
    "ステータス",
]

EXPECTED_COLUMNS_FIAT_DEPOSIT_JP = [
    "時間",
    "方法",
    "入金額",
    "受取金額",
    "手数料",
    "ステータス",
    "取引ID",
]

EXPECTED_COLUMNS_FIAT_WITHDRAW_JP = [
    "時間",
    "方法",
    "出金額",
    "受取金額",
    "手数料",
    "ステータス",
    "取引ID",
]

EXPECTED_COLUMNS_FIAT_CONVERSION_JP = [
    "方法",
    "金額",
    "価格",
    "最終金額",
    "作成日時",
    "ステータス",
    "取引ID",
]

HEADER_SCAN_MAX_ROWS = 64
NO_DATA_TEXT = "条件に一致するデータがありません。"

IGNORED_INTERNAL_OPERATIONS = {
    "Transfer Between Main and Funding Wallet",
    "Transfer Between Spot and Strategy Account",
    "Transfer Between Main Account and Sub-Account",
    "Simple Earn Locked Subscription",
    "Simple Earn Locked Redemption",
}

REWARD_OPERATIONS = {
    "Simple Earn Locked Rewards",
    "Launchpool Airdrop - JP Users Distribution",
}

SINGLE_ROW_OPERATIONS = {
    "Deposit",
    "Fiat Withdraw",
    "Buy Crypto With Fiat",
    "Sell Crypto To Fiat",
} | REWARD_OPERATIONS

TRADE_FAMILY_OPERATIONS = {
    "Transaction Fee",
    "Transaction Buy",
    "Transaction Spend",
    "Transaction Revenue",
    "Transaction Sold",
}


class BinanceJapanParser(BaseParser):
    exchange_name = "binance_japan"

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}

    def parse(self, path: Path) -> ImportBatchResult:
        rows, detected_layout, header_row_number = self._load_rows(path)
        if detected_layout == "csv_spot_trade_history":
            return self._parse_spot_trade_history(
                path,
                rows,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if detected_layout == "csv_deposit_history":
            return self._parse_transfer_history(
                path,
                rows,
                tx_type=TransactionType.TRANSFER_IN,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if detected_layout == "csv_withdraw_history":
            return self._parse_transfer_history(
                path,
                rows,
                tx_type=TransactionType.TRANSFER_OUT,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if detected_layout == "csv_fiat_deposit_history":
            return self._parse_fiat_history(
                path,
                rows,
                tx_type=TransactionType.TRANSFER_IN,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if detected_layout == "csv_fiat_withdraw_history":
            return self._parse_fiat_history(
                path,
                rows,
                tx_type=TransactionType.TRANSFER_OUT,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if detected_layout == "csv_fiat_conversion_history":
            return self._parse_fiat_conversion_history(
                path,
                rows,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        if rows and self._is_japanese_balance_history_schema(rows[0][1]):
            return self._parse_japanese_balance_history(
                path,
                rows,
                detected_layout=detected_layout,
                header_row_number=header_row_number,
            )
        return self._parse_trade_export(
            path,
            rows,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _batch_id(self, path: Path) -> str:
        return f"binance_japan_{safe_slug(path.stem)}_{int(path.stat().st_mtime)}"

    def _source_kind(self, path: Path) -> ImportSourceKind:
        return ImportSourceKind.XLSX if path.suffix.lower() != ".csv" else ImportSourceKind.CSV

    def _load_rows(self, path: Path) -> tuple[list[tuple[int, dict[str, Any]]], str | None, int | None]:
        if path.suffix.lower() == ".csv":
            rows: list[tuple[int, dict[str, Any]]] = []
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.DictReader(fh)
                detected_layout = self._detect_csv_layout(reader.fieldnames or [])
                for idx, row in enumerate(reader, start=2):
                    rows.append((idx, {k: v for k, v in row.items()}))
            if detected_layout is None:
                detected_layout = "csv_japanese_balance_history" if rows and self._is_japanese_balance_history_schema(rows[0][1]) else "csv_trade_export"
            return rows, detected_layout, 1

        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row_number, headers, detected_layout = self._detect_xlsx_header(ws)
        if not headers:
            wb.close()
            wb = load_workbook(path, read_only=False, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row_number, headers, detected_layout = self._detect_xlsx_header(ws)
        if not headers:
            return [], "xlsx_unknown_layout", None

        active_columns = [(index, header) for index, header in enumerate(headers) if header]
        xlsx_rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, cells in enumerate(ws.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
            if not active_columns:
                continue
            if not any(index < len(cells) and cells[index] not in (None, "") for index, _ in active_columns):
                continue
            row = {
                header: cells[index] if index < len(cells) else None
                for index, header in active_columns
            }
            xlsx_rows.append((row_number, row))
        return xlsx_rows, detected_layout, header_row_number

    def _detect_xlsx_header(self, worksheet: Any) -> tuple[int, list[str], str]:
        fallback_header: list[str] | None = None
        fallback_row_number: int | None = None

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(getattr(worksheet, "max_row") or HEADER_SCAN_MAX_ROWS, HEADER_SCAN_MAX_ROWS),
                values_only=True,
            ),
            start=1,
        ):
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            if fallback_header is None and any(headers):
                fallback_header = headers
                fallback_row_number = row_number
            key_set = {header for header in headers if header}
            if set(EXPECTED_COLUMNS_JP).issubset(key_set):
                return row_number, headers, "xlsx_japanese_balance_history_report"
            if set(EXPECTED_COLUMNS).issubset(key_set):
                return row_number, headers, "xlsx_trade_export"

        if fallback_header is not None and fallback_row_number is not None:
            return fallback_row_number, fallback_header, "xlsx_unrecognized_header"
        return 1, [], "xlsx_unknown_layout"

    def _detect_csv_layout(self, fieldnames: Sequence[str]) -> str | None:
        key_set = {name.strip() for name in fieldnames if name}
        if set(EXPECTED_COLUMNS_JP).issubset(key_set):
            return "csv_japanese_balance_history"
        if set(EXPECTED_COLUMNS_SPOT_TRADE_JP).issubset(key_set):
            return "csv_spot_trade_history"
        if set(EXPECTED_COLUMNS_FIAT_WITHDRAW_JP).issubset(key_set):
            return "csv_fiat_withdraw_history"
        if set(EXPECTED_COLUMNS_FIAT_DEPOSIT_JP).issubset(key_set):
            return "csv_fiat_deposit_history"
        if set(EXPECTED_COLUMNS_FIAT_CONVERSION_JP).issubset(key_set):
            return "csv_fiat_conversion_history"
        if set(EXPECTED_COLUMNS_WITHDRAW_JP).issubset(key_set):
            return "csv_withdraw_history"
        if set(EXPECTED_COLUMNS_DEPOSIT_JP).issubset(key_set):
            return "csv_deposit_history"
        if set(EXPECTED_COLUMNS).issubset(key_set):
            return "csv_trade_export"
        return None

    def _parse_trade_export(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        unknown_columns: set[str] = set()
        unknown_types: set[str] = set()
        for row_number, row in rows:
            tx = self._row_to_trade_tx(path, row_number, row)
            unknown_columns.update(tx.raw_payload.get("_unknown_columns", []))
            if tx.tx_type is TransactionType.UNKNOWN:
                unknown_types.add(str(row.get("Type", "")))
            transactions.append(tx)

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=sorted(unknown_columns),
            unknown_tx_types=sorted(filter(None, unknown_types)),
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _parse_spot_trade_history(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        for row_number, row in rows:
            if self._is_no_data_row(row):
                continue
            tx = self._row_to_spot_trade_tx(path, row_number, row)
            transactions.append(tx)

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=[],
            unknown_tx_types=[],
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _parse_transfer_history(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        tx_type: TransactionType,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        for row_number, row in rows:
            if self._is_no_data_row(row):
                continue
            tx = self._row_to_transfer_tx(path, row_number, row, tx_type=tx_type)
            transactions.append(tx)

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=[],
            unknown_tx_types=[],
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _parse_fiat_history(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        tx_type: TransactionType,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        for row_number, row in rows:
            if self._is_no_data_row(row):
                continue
            tx = self._row_to_fiat_transfer_tx(path, row_number, row, tx_type=tx_type)
            transactions.append(tx)

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=[],
            unknown_tx_types=[],
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _parse_fiat_conversion_history(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        for row_number, row in rows:
            if self._is_no_data_row(row):
                continue
            transactions.append(self._row_to_unknown_fiat_conversion_tx(path, row_number, row))

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=[],
            unknown_tx_types=[],
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _parse_japanese_balance_history(
        self,
        path: Path,
        rows: list[tuple[int, dict[str, Any]]],
        *,
        detected_layout: str | None = None,
        header_row_number: int | None = None,
    ) -> ImportBatchResult:
        transactions: list[NormalizedTransaction] = []
        unknown_columns: set[str] = set()
        unknown_types: set[str] = set()
        index = 0
        while index < len(rows):
            row_number, row = rows[index]
            unknown_columns.update(key for key in row.keys() if key and key not in EXPECTED_COLUMNS_JP)
            operation = str(row.get("操作") or "").strip()

            if operation in IGNORED_INTERNAL_OPERATIONS:
                index += 1
                continue

            if operation == "Binance Convert":
                group, index = self._collect_nearby_group(rows, index, operation=operation, seconds=2)
                transactions.extend(self._parse_simple_pair_group(path, group, operation))
                continue

            if operation in TRADE_FAMILY_OPERATIONS:
                group, index = self._collect_same_timestamp_group(rows, index, allowed_operations=TRADE_FAMILY_OPERATIONS)
                transactions.extend(self._parse_transaction_family_group(path, group))
                continue

            if operation == "Small Assets Exchange BNB":
                group, index = self._collect_same_timestamp_group(rows, index, allowed_operations={operation})
                transactions.extend(self._parse_small_assets_group(path, group))
                continue

            if operation == "Strategy Trading Fee Rebate":
                group, index = self._collect_same_timestamp_group(rows, index, allowed_operations={operation})
                transactions.extend(self._parse_rebate_group(path, group))
                continue

            if operation in SINGLE_ROW_OPERATIONS:
                tx = self._parse_single_japanese_row(path, row_number, row)
                if tx is not None:
                    transactions.append(tx)
                index += 1
                continue

            unknown_types.add(operation)
            tx = self._make_unknown_japanese_tx(path, row_number, row, reason="未知の取引種別")
            transactions.append(tx)
            index += 1

        batch_id = self._batch_id(path)
        return ImportBatchResult(
            batch_id=batch_id,
            source_file=path.name,
            source_kind=self._source_kind(path),
            transaction_count=len(transactions),
            review_required_count=sum(1 for tx in transactions if tx.review_flag),
            duplicate_count=0,
            unknown_column_names=sorted(unknown_columns),
            unknown_tx_types=sorted(filter(None, unknown_types)),
            transactions=transactions,
            detected_layout=detected_layout,
            header_row_number=header_row_number,
        )

    def _row_to_trade_tx(self, path: Path, row_number: int, row: dict[str, Any]) -> NormalizedTransaction:
        unknown_columns = sorted(key for key in row.keys() if key and key not in EXPECTED_COLUMNS)
        timestamp_utc = parse_utc_timestamp(str(row.get("Date(UTC)") or ""))
        timestamp_jst = utc_to_jst(timestamp_utc)
        base_asset = (row.get("Base Asset") or "").strip().upper() or None
        quote_asset = (row.get("Quote Asset") or "").strip().upper() or None
        pair = (row.get("Pair") or "").strip()
        order_type = (str(row.get("Type") or "")).strip().upper()
        quantity = to_decimal(row.get("Amount"))
        total = to_decimal(row.get("Total"))
        price = to_decimal(row.get("Price"))
        fee_amount = to_decimal(row.get("Fee"))
        fee_asset = (row.get("Fee Coin") or "").strip().upper() or None

        tx_type, side = self._classify_trade_export(order_type=order_type, quote_asset=quote_asset)
        price_per_unit_jpy = price if quote_asset == "JPY" else None
        gross_amount_jpy = total if quote_asset == "JPY" else None
        fee_jpy = fee_amount if fee_asset == "JPY" else None

        review_reasons: list[str] = []
        if quote_asset != "JPY":
            review_reasons.append("JPY換算列がないため要確認")
        if tx_type is TransactionType.UNKNOWN:
            review_reasons.append("未知の取引種別")
        if fee_amount and fee_asset and fee_asset != "JPY":
            review_reasons.append("手数料のJPY換算が未確定")

        raw_payload = {**row, "_unknown_columns": unknown_columns}
        tx_hash = hashlib.sha1(
            f"{path.name}:{row_number}:{pair}:{timestamp_utc}:{order_type}".encode("utf-8")
        ).hexdigest()[:16]
        return NormalizedTransaction(
            id=f"tx_{tx_hash}",
            source_exchange=self.exchange_name,
            source_file=path.name,
            raw_row_number=row_number,
            timestamp_jst=timestamp_jst,
            timestamp_utc=timestamp_utc,
            tx_type=tx_type,
            base_asset=base_asset,
            quote_asset=quote_asset,
            quantity=quantity,
            quote_quantity=total if quote_asset and quote_asset != "JPY" else None,
            unit_price_quote=price,
            price_per_unit_jpy=price_per_unit_jpy,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_jpy,
            side=side,
            note=f"pair={pair}; order_type={order_type}",
            raw_payload=raw_payload,
            classification_status=(
                ClassificationStatus.CLASSIFIED if not review_reasons else ClassificationStatus.REVIEW_REQUIRED
            ),
            review_flag=bool(review_reasons),
            review_reasons=review_reasons,
            source_kind=self._source_kind(path),
            jpy_rate_source="file:quote_jpy" if quote_asset == "JPY" else None,
        )

    def _row_to_spot_trade_tx(self, path: Path, row_number: int, row: dict[str, Any]) -> NormalizedTransaction:
        timestamp_utc = parse_utc_timestamp(str(row.get("時間") or ""))
        if timestamp_utc is None:
            timestamp_jst = self._parse_japanese_timestamp(str(row.get("時間") or ""))
            timestamp_utc = timestamp_jst.astimezone(timezone.utc) if timestamp_jst is not None else None
        else:
            timestamp_jst = utc_to_jst(timestamp_utc)
        executed_quantity, base_asset = self._parse_amount_with_asset(row.get("実行済み"))
        total_quantity, quote_asset = self._parse_amount_with_asset(row.get("金額"))
        fee_amount, fee_asset = self._parse_amount_with_asset(row.get("手数料"))
        side_text = str(row.get("サイド") or "").strip().upper()
        pair = str(row.get("ペア") or "").strip().upper()
        price = to_decimal(row.get("価格"))

        review_reasons: list[str] = []
        if executed_quantity is None or base_asset is None:
            review_reasons.append("約定数量の解析に失敗")
        if total_quantity is None or quote_asset is None:
            review_reasons.append("金額列の解析に失敗")

        side = Side.BUY if side_text == "BUY" else Side.SELL if side_text == "SELL" else Side.NONE
        if side is Side.NONE:
            review_reasons.append("売買サイドの解析に失敗")

        if quote_asset == "JPY":
            tx_type = TransactionType.BUY if side is Side.BUY else TransactionType.SELL if side is Side.SELL else TransactionType.UNKNOWN
            price_per_unit_jpy = price
            gross_amount_jpy = total_quantity
            quote_quantity = None
            jpy_rate_source = "file:spot_trade_jpy"
        else:
            tx_type = TransactionType.CRYPTO_SWAP if side in {Side.BUY, Side.SELL} else TransactionType.UNKNOWN
            price_per_unit_jpy = None
            gross_amount_jpy = None
            quote_quantity = total_quantity
            jpy_rate_source = None
            review_reasons.append("JPY換算列がないため要確認")

        fee_jpy = fee_amount if fee_asset == "JPY" else None
        if fee_amount and fee_asset and fee_asset != "JPY":
            review_reasons.append("手数料のJPY換算が未確定")

        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=tx_type,
            base_asset=base_asset,
            quote_asset=quote_asset,
            quantity=executed_quantity,
            quote_quantity=quote_quantity,
            unit_price_quote=price,
            price_per_unit_jpy=price_per_unit_jpy,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_jpy,
            side=side,
            note=f"pair={pair}; source=spot_trade_history",
            raw_payload={"source_rows": [row]},
            review_reasons=review_reasons,
            jpy_rate_source=jpy_rate_source,
        )

    def _row_to_transfer_tx(
        self,
        path: Path,
        row_number: int,
        row: dict[str, Any],
        *,
        tx_type: TransactionType,
    ) -> NormalizedTransaction:
        timestamp_utc = parse_utc_timestamp(str(row.get("時間") or ""))
        if timestamp_utc is None:
            timestamp_jst = self._parse_japanese_timestamp(str(row.get("時間") or ""))
            timestamp_utc = timestamp_jst.astimezone(timezone.utc) if timestamp_jst is not None else None
        else:
            timestamp_jst = utc_to_jst(timestamp_utc)
        asset = (row.get("コイン") or "").strip().upper() or None
        quantity = to_decimal(row.get("金額"))
        fee_amount = to_decimal(row.get("手数料"))
        fee_asset = asset if fee_amount is not None and fee_amount != 0 else None
        review_reasons: list[str] = []
        if asset is None:
            review_reasons.append("コイン列の解析に失敗")
        if quantity is None:
            review_reasons.append("数量列の解析に失敗")
        gross_amount_jpy = quantity if asset == "JPY" else None
        fee_jpy = fee_amount if fee_asset == "JPY" else None
        if fee_amount and fee_asset and fee_asset != "JPY":
            review_reasons.append("手数料のJPY換算が未確定")
        note_parts = [f"network={row.get('ネットワーク') or ''}", f"status={row.get('ステータス') or ''}"]
        txid = str(row.get("トランザクションID") or "").strip()
        if txid:
            note_parts.append(f"txid={txid}")
        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=tx_type,
            base_asset=asset,
            quote_asset=None,
            quantity=quantity,
            quote_quantity=None,
            unit_price_quote=None,
            price_per_unit_jpy=Decimal("1") if asset == "JPY" and quantity is not None else None,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_jpy,
            side=Side.NONE,
            note="; ".join(part for part in note_parts if part),
            raw_payload={"source_rows": [row]},
            review_reasons=review_reasons,
            jpy_rate_source="file:transfer_jpy" if gross_amount_jpy is not None else None,
        )

    def _row_to_fiat_transfer_tx(
        self,
        path: Path,
        row_number: int,
        row: dict[str, Any],
        *,
        tx_type: TransactionType,
    ) -> NormalizedTransaction:
        timestamp_jst = self._parse_japanese_timestamp(str(row.get("時間") or ""))
        timestamp_utc = timestamp_jst.astimezone(timezone.utc) if timestamp_jst is not None else None
        gross_column = "入金額" if tx_type is TransactionType.TRANSFER_IN else "出金額"
        gross_amount, gross_asset = self._parse_amount_with_asset(row.get(gross_column))
        net_amount, net_asset = self._parse_amount_with_asset(row.get("受取金額"))
        fee_amount, fee_asset = self._parse_amount_with_asset(row.get("手数料"))
        review_reasons: list[str] = []
        asset = gross_asset or net_asset or "JPY"
        quantity = net_amount if tx_type is TransactionType.TRANSFER_IN else gross_amount
        if asset != "JPY":
            review_reasons.append("法定通貨履歴なのに JPY 以外が含まれています")
        if quantity is None:
            review_reasons.append("数量列の解析に失敗")
        if fee_amount and fee_asset and fee_asset != "JPY":
            review_reasons.append("手数料のJPY換算が未確定")
        note_parts = [
            f"method={row.get('方法') or ''}",
            f"status={row.get('ステータス') or ''}",
        ]
        txid = str(row.get("取引ID") or "").strip()
        if txid:
            note_parts.append(f"txid={txid}")
        if net_amount is not None and tx_type is TransactionType.TRANSFER_OUT:
            note_parts.append(f"net_received_jpy={net_amount}")
        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=tx_type,
            base_asset=asset,
            quote_asset=None,
            quantity=quantity,
            quote_quantity=None,
            unit_price_quote=None,
            price_per_unit_jpy=Decimal("1") if quantity is not None else None,
            gross_amount_jpy=quantity,
            fee_asset="JPY" if fee_amount is not None else fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_amount,
            side=Side.NONE,
            note="; ".join(part for part in note_parts if part),
            raw_payload={"source_rows": [row]},
            review_reasons=review_reasons,
            jpy_rate_source="file:fiat_history_jpy",
        )

    def _row_to_unknown_fiat_conversion_tx(
        self,
        path: Path,
        row_number: int,
        row: dict[str, Any],
    ) -> NormalizedTransaction:
        timestamp_jst = self._parse_japanese_timestamp(str(row.get("作成日時") or ""))
        timestamp_utc = timestamp_jst.astimezone(timezone.utc) if timestamp_jst is not None else None
        amount, asset = self._parse_amount_with_asset(row.get("金額"))
        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=TransactionType.UNKNOWN,
            base_asset=asset,
            quote_asset=None,
            quantity=amount,
            quote_quantity=None,
            unit_price_quote=None,
            price_per_unit_jpy=None,
            gross_amount_jpy=None,
            fee_asset=None,
            fee_amount=None,
            fee_jpy=None,
            side=Side.NONE,
            note=f"fiat_method={row.get('方法') or ''}",
            raw_payload={"source_rows": [row]},
            review_reasons=["法定通貨による交換履歴の自動解釈は未対応のため要確認"],
            jpy_rate_source=None,
        )

    def _is_japanese_balance_history_schema(self, row: dict[str, Any]) -> bool:
        keys = set(key for key in row.keys() if key)
        return {"時間", "操作", "コイン", "変更"}.issubset(keys)

    def _classify_trade_export(
        self,
        *,
        order_type: str,
        quote_asset: str | None,
    ) -> tuple[TransactionType, Side]:
        if order_type == "BUY":
            if quote_asset == "JPY":
                return TransactionType.BUY, Side.BUY
            return TransactionType.CRYPTO_SWAP, Side.BUY
        if order_type == "SELL":
            if quote_asset == "JPY":
                return TransactionType.SELL, Side.SELL
            return TransactionType.CRYPTO_SWAP, Side.SELL
        if "DEPOSIT" in order_type:
            return TransactionType.TRANSFER_IN, Side.NONE
        if "WITHDRAW" in order_type:
            return TransactionType.TRANSFER_OUT, Side.NONE
        if any(token in order_type for token in ("STAK", "REWARD", "BONUS", "EARN")):
            return TransactionType.REWARD, Side.NONE
        return TransactionType.UNKNOWN, Side.NONE

    def _collect_nearby_group(
        self,
        rows: list[tuple[int, dict[str, Any]]],
        start_index: int,
        *,
        operation: str,
        seconds: int,
    ) -> tuple[list[tuple[int, dict[str, Any]]], int]:
        group = [rows[start_index]]
        base_row = rows[start_index][1]
        base_account = str(base_row.get("アカウント") or "")
        base_time = self._parse_japanese_timestamp(str(base_row.get("時間") or ""))
        next_index = start_index + 1
        while next_index < len(rows):
            _, candidate = rows[next_index]
            candidate_time = self._parse_japanese_timestamp(str(candidate.get("時間") or ""))
            if base_time is None or candidate_time is None:
                break
            if (candidate_time - base_time).total_seconds() > seconds:
                break
            if (
                str(candidate.get("操作") or "") == operation
                and str(candidate.get("アカウント") or "") == base_account
            ):
                group.append(rows[next_index])
            next_index += 1
        return group, next_index

    def _collect_same_timestamp_group(
        self,
        rows: list[tuple[int, dict[str, Any]]],
        start_index: int,
        *,
        allowed_operations: set[str],
    ) -> tuple[list[tuple[int, dict[str, Any]]], int]:
        group: list[tuple[int, dict[str, Any]]] = []
        base_row = rows[start_index][1]
        base_account = str(base_row.get("アカウント") or "")
        base_time = str(base_row.get("時間") or "")
        next_index = start_index + 1
        if str(base_row.get("操作") or "") in allowed_operations:
            group.append(rows[start_index])
        while next_index < len(rows):
            _, candidate = rows[next_index]
            if str(candidate.get("時間") or "") != base_time:
                break
            if (
                str(candidate.get("アカウント") or "") == base_account
                and str(candidate.get("操作") or "") in allowed_operations
            ):
                group.append(rows[next_index])
            next_index += 1
        return group, next_index

    def _parse_transaction_family_group(
        self,
        path: Path,
        group: list[tuple[int, dict[str, Any]]],
    ) -> list[NormalizedTransaction]:
        buy_rows = [item for item in group if item[1].get("操作") == "Transaction Buy"]
        spend_rows = [item for item in group if item[1].get("操作") == "Transaction Spend"]
        sold_rows = [item for item in group if item[1].get("操作") == "Transaction Sold"]
        revenue_rows = [item for item in group if item[1].get("操作") == "Transaction Revenue"]
        fee_rows = [item for item in group if item[1].get("操作") == "Transaction Fee"]

        transactions: list[NormalizedTransaction] = []
        used_indexes: set[int] = set()

        if buy_rows and spend_rows:
            pair_count = min(len(buy_rows), len(spend_rows))
            for idx in range(pair_count):
                fee_row = fee_rows[idx] if idx < len(fee_rows) else None
                if fee_row is not None:
                    used_indexes.add(group.index(fee_row))
                transactions.append(
                    self._make_tx_from_pair_rows(
                        path,
                        positive_row=buy_rows[idx],
                        negative_row=spend_rows[idx],
                        fee_row=fee_row,
                        operation="Transaction Trade",
                    )
                )
                used_indexes.add(group.index(buy_rows[idx]))
                used_indexes.add(group.index(spend_rows[idx]))

        if sold_rows and revenue_rows:
            pair_count = min(len(sold_rows), len(revenue_rows))
            fee_offset = min(len(buy_rows), len(spend_rows))
            for idx in range(pair_count):
                fee_index = fee_offset + idx
                fee_row = fee_rows[fee_index] if fee_index < len(fee_rows) else None
                if fee_row is not None:
                    used_indexes.add(group.index(fee_row))
                transactions.append(
                    self._make_tx_from_pair_rows(
                        path,
                        positive_row=revenue_rows[idx],
                        negative_row=sold_rows[idx],
                        fee_row=fee_row,
                        operation="Transaction Trade",
                    )
                )
                used_indexes.add(group.index(revenue_rows[idx]))
                used_indexes.add(group.index(sold_rows[idx]))

        if transactions:
            for idx, (row_number, row) in enumerate(group):
                if idx not in used_indexes:
                    transactions.append(
                        self._make_unknown_japanese_tx(path, row_number, row, reason="取引グループの一部が余りました")
                    )
            return transactions

        for row_number, row in group:
            transactions.append(
                self._make_unknown_japanese_tx(path, row_number, row, reason="取引グループの解釈に失敗")
            )
        return transactions

    def _parse_simple_pair_group(
        self,
        path: Path,
        group: list[tuple[int, dict[str, Any]]],
        operation: str,
    ) -> list[NormalizedTransaction]:
        positives = [item for item in group if self._jp_change(item[1]) and self._jp_change(item[1]) > 0]
        negatives = [item for item in group if self._jp_change(item[1]) and self._jp_change(item[1]) < 0]
        transactions: list[NormalizedTransaction] = []
        while positives and negatives:
            transactions.append(
                self._make_tx_from_pair_rows(
                    path,
                    positive_row=positives.pop(0),
                    negative_row=negatives.pop(0),
                    fee_row=None,
                    operation=operation,
                )
            )
        for row_number, row in positives + negatives:
            transactions.append(self._make_unknown_japanese_tx(path, row_number, row, reason="取引ペアの解釈に失敗"))
        return transactions

    def _parse_small_assets_group(
        self,
        path: Path,
        group: list[tuple[int, dict[str, Any]]],
    ) -> list[NormalizedTransaction]:
        grouped_by_note: dict[str, list[tuple[int, dict[str, Any]]]] = {}
        for item in group:
            note = str(item[1].get("備考") or "").strip()
            grouped_by_note.setdefault(note, []).append(item)

        transactions: list[NormalizedTransaction] = []
        for note, note_rows in grouped_by_note.items():
            transactions.extend(self._parse_simple_pair_group(path, note_rows, f"Small Assets Exchange BNB:{note or 'no_note'}"))
        return transactions

    def _parse_rebate_group(
        self,
        path: Path,
        group: list[tuple[int, dict[str, Any]]],
    ) -> list[NormalizedTransaction]:
        positives = [item for item in group if self._jp_change(item[1]) and self._jp_change(item[1]) > 0]
        negatives = [item for item in group if self._jp_change(item[1]) and self._jp_change(item[1]) < 0]
        transactions: list[NormalizedTransaction] = []
        while positives and negatives:
            transactions.append(
                self._make_tx_from_pair_rows(
                    path,
                    positive_row=positives.pop(0),
                    negative_row=negatives.pop(0),
                    fee_row=None,
                    operation="Strategy Trading Fee Rebate",
                )
            )
        for row_number, row in positives:
            tx = self._parse_single_japanese_row(path, row_number, row, force_reward=True)
            if tx is not None:
                transactions.append(tx)
        for row_number, row in negatives:
            transactions.append(self._make_unknown_japanese_tx(path, row_number, row, reason="リベート差分の解釈に失敗"))
        return transactions

    def _parse_single_japanese_row(
        self,
        path: Path,
        row_number: int,
        row: dict[str, Any],
        *,
        force_reward: bool = False,
    ) -> NormalizedTransaction | None:
        operation = str(row.get("操作") or "").strip()
        asset = (row.get("コイン") or "").strip().upper() or None
        change = self._jp_change(row)
        if change is None or asset is None or change == 0:
            return self._make_unknown_japanese_tx(path, row_number, row, reason="数量が読み取れません")

        timestamp_utc, timestamp_jst = self._japanese_row_timestamps(row)
        note = f"ledger_op={operation}"
        review_reasons: list[str] = []
        tx_type = TransactionType.UNKNOWN
        side = Side.NONE
        quantity = abs(change)
        quote_asset = None
        gross_amount_jpy = None
        price_per_unit_jpy = None

        if force_reward or operation in REWARD_OPERATIONS:
            tx_type = TransactionType.REWARD
            if asset == "JPY":
                gross_amount_jpy = quantity
                price_per_unit_jpy = 1
            else:
                review_reasons.append("JPY換算列がないため要確認")
        elif operation == "Deposit":
            tx_type = TransactionType.TRANSFER_IN
            if change < 0:
                review_reasons.append("Deposit なのに数量がマイナスです")
        elif operation == "Fiat Withdraw":
            tx_type = TransactionType.TRANSFER_OUT
            if change > 0:
                review_reasons.append("Withdraw なのに数量がプラスです")
        elif operation == "Buy Crypto With Fiat":
            tx_type = TransactionType.BUY
            side = Side.BUY
            quote_asset = "JPY"
            review_reasons.append("法定通貨側の変動行がCSVにないため要確認")
        elif operation == "Sell Crypto To Fiat":
            tx_type = TransactionType.SELL
            side = Side.SELL
            quote_asset = "JPY"
            review_reasons.append("法定通貨側の変動行がCSVにないため要確認")
        else:
            return self._make_unknown_japanese_tx(path, row_number, row, reason="単独行の解釈に失敗")

        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=tx_type,
            base_asset=asset,
            quote_asset=quote_asset,
            quantity=quantity,
            quote_quantity=None,
            unit_price_quote=None,
            price_per_unit_jpy=price_per_unit_jpy,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=None,
            fee_amount=None,
            fee_jpy=None,
            side=side,
            note=note,
            raw_payload={"source_rows": [row]},
            review_reasons=review_reasons,
            jpy_rate_source="file:journal_jpy" if gross_amount_jpy is not None else None,
        )

    def _make_tx_from_pair_rows(
        self,
        path: Path,
        *,
        positive_row: tuple[int, dict[str, Any]],
        negative_row: tuple[int, dict[str, Any]],
        fee_row: tuple[int, dict[str, Any]] | None,
        operation: str,
    ) -> NormalizedTransaction:
        pos_row_number, pos_row = positive_row
        neg_row_number, neg_row = negative_row
        fee_asset = None
        fee_amount = None
        fee_jpy = None
        if fee_row is not None:
            _, fee_data = fee_row
            fee_asset = (fee_data.get("コイン") or "").strip().upper() or None
            fee_amount = abs(self._jp_change(fee_data) or 0)
            fee_jpy = fee_amount if fee_asset == "JPY" else None

        pos_asset = (pos_row.get("コイン") or "").strip().upper() or None
        neg_asset = (neg_row.get("コイン") or "").strip().upper() or None
        pos_amount = abs(self._jp_change(pos_row) or 0)
        neg_amount = abs(self._jp_change(neg_row) or 0)
        timestamp_utc, timestamp_jst = self._japanese_row_timestamps(pos_row)
        review_reasons: list[str] = []

        tx_type = TransactionType.CRYPTO_SWAP
        side = Side.BUY
        base_asset = pos_asset
        quote_asset = neg_asset
        quantity = pos_amount
        quote_quantity = neg_amount
        unit_price_quote = None if pos_amount in (None, 0) else neg_amount / pos_amount
        price_per_unit_jpy = None
        gross_amount_jpy = None
        jpy_rate_source = None

        if pos_asset == "JPY" and neg_asset:
            tx_type = TransactionType.SELL
            side = Side.SELL
            base_asset = neg_asset
            quote_asset = "JPY"
            quantity = neg_amount
            quote_quantity = None
            unit_price_quote = None if neg_amount in (None, 0) else pos_amount / neg_amount
            price_per_unit_jpy = unit_price_quote
            gross_amount_jpy = pos_amount
            jpy_rate_source = "file:journal_jpy"
        elif neg_asset == "JPY" and pos_asset:
            tx_type = TransactionType.BUY
            side = Side.BUY
            base_asset = pos_asset
            quote_asset = "JPY"
            quantity = pos_amount
            quote_quantity = None
            unit_price_quote = None if pos_amount in (None, 0) else neg_amount / pos_amount
            price_per_unit_jpy = unit_price_quote
            gross_amount_jpy = neg_amount
            jpy_rate_source = "file:journal_jpy"
        else:
            review_reasons.append("JPY換算列がないため要確認")

        raw_payload = {
            "source_rows": [pos_row, neg_row] + ([fee_row[1]] if fee_row else []),
            "operation": operation,
        }
        return self._build_transaction(
            path=path,
            row_number=min(pos_row_number, neg_row_number),
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=tx_type,
            base_asset=base_asset,
            quote_asset=quote_asset,
            quantity=quantity,
            quote_quantity=quote_quantity,
            unit_price_quote=unit_price_quote,
            price_per_unit_jpy=price_per_unit_jpy,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_jpy,
            side=side,
            note=f"ledger_op={operation}",
            raw_payload=raw_payload,
            review_reasons=review_reasons,
            jpy_rate_source=jpy_rate_source,
        )

    def _make_unknown_japanese_tx(
        self,
        path: Path,
        row_number: int,
        row: dict[str, Any],
        *,
        reason: str,
    ) -> NormalizedTransaction:
        timestamp_utc, timestamp_jst = self._japanese_row_timestamps(row)
        asset = (row.get("コイン") or "").strip().upper() or None
        change = self._jp_change(row)
        quantity = abs(change) if change is not None else None
        return self._build_transaction(
            path=path,
            row_number=row_number,
            timestamp_utc=timestamp_utc,
            timestamp_jst=timestamp_jst,
            tx_type=TransactionType.UNKNOWN,
            base_asset=asset,
            quote_asset=None,
            quantity=quantity,
            quote_quantity=None,
            unit_price_quote=None,
            price_per_unit_jpy=None,
            gross_amount_jpy=None,
            fee_asset=None,
            fee_amount=None,
            fee_jpy=None,
            side=Side.NONE,
            note=f"ledger_op={row.get('操作')}",
            raw_payload={"source_rows": [row]},
            review_reasons=[reason],
            jpy_rate_source=None,
        )

    def _build_transaction(
        self,
        *,
        path: Path,
        row_number: int,
        timestamp_utc: datetime | None,
        timestamp_jst: datetime | None,
        tx_type: TransactionType,
        base_asset: str | None,
        quote_asset: str | None,
        quantity,
        quote_quantity,
        unit_price_quote,
        price_per_unit_jpy,
        gross_amount_jpy,
        fee_asset: str | None,
        fee_amount,
        fee_jpy,
        side: Side,
        note: str,
        raw_payload: dict[str, Any],
        review_reasons: list[str],
        jpy_rate_source: str | None,
    ) -> NormalizedTransaction:
        digest = hashlib.sha1(
            f"{path.name}:{row_number}:{timestamp_utc}:{tx_type.value}:{base_asset}:{quote_asset}:{note}".encode("utf-8")
        ).hexdigest()[:16]
        return NormalizedTransaction(
            id=f"tx_{digest}",
            source_exchange=self.exchange_name,
            source_file=path.name,
            raw_row_number=row_number,
            timestamp_jst=timestamp_jst,
            timestamp_utc=timestamp_utc,
            tx_type=tx_type,
            base_asset=base_asset,
            quote_asset=quote_asset,
            quantity=quantity,
            quote_quantity=quote_quantity,
            unit_price_quote=unit_price_quote,
            price_per_unit_jpy=price_per_unit_jpy,
            gross_amount_jpy=gross_amount_jpy,
            fee_asset=fee_asset,
            fee_amount=fee_amount,
            fee_jpy=fee_jpy,
            side=side,
            note=note,
            raw_payload=raw_payload,
            classification_status=(
                ClassificationStatus.CLASSIFIED if not review_reasons else ClassificationStatus.REVIEW_REQUIRED
            ),
            review_flag=bool(review_reasons),
            review_reasons=review_reasons,
            source_kind=self._source_kind(path),
            jpy_rate_source=jpy_rate_source,
        )

    def _jp_change(self, row: dict[str, Any]):
        return to_decimal(row.get("変更"))

    def _parse_amount_with_asset(self, value: Any) -> tuple[Any, str | None]:
        text = str(value or "").strip()
        if not text:
            return None, None
        match = re.match(r"^\s*([+-]?\d+(?:\.\d+)?)\s*([A-Za-z]+)\s*$", text)
        if match:
            return to_decimal(match.group(1)), match.group(2).upper()
        decimal_value = to_decimal(text)
        return decimal_value, None

    def _is_no_data_row(self, row: dict[str, Any]) -> bool:
        return any(NO_DATA_TEXT in str(value or "") for value in row.values())

    def _parse_japanese_timestamp(self, value: str) -> datetime | None:
        text = value.strip()
        if not text:
            return None
        for fmt in ("%y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).replace(tzinfo=JST)
            except ValueError:
                continue
        return None

    def _japanese_row_timestamps(self, row: dict[str, Any]) -> tuple[datetime | None, datetime | None]:
        local_dt = self._parse_japanese_timestamp(str(row.get("時間") or ""))
        if local_dt is None:
            return None, None
        return local_dt.astimezone(timezone.utc), local_dt
